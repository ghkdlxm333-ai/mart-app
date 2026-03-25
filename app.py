import streamlit as st
import pandas as pd
import io
import os
import openpyxl
from datetime import datetime, timedelta, timezone

# --- 1. 페이지 설정 ---
st.set_page_config(page_title="Tesco 수주 자동 변환기", layout="wide")

# 고정된 서식 파일 이름
MASTER_TEMPLATE = "Tesco_서식파일_업데이트용.xlsx"

# 최종 통합 수주업로드 양식 컬럼 (공백 포함)
FINAL_COLUMNS = [
    '출고구분', '수주일자', '납품일자', '발주처코드', '발주처', 
    '배송코드', '배송지', '상품코드', '상품명', 'UNIT수량', 
    'UNIT단가', '금        액', '부  가   세', 'LOT', '특이사항1', 'Type', '특이사항2'
]

EXPORT_HEADERS = [
    '출고구분', '수주일자', '납품일자', '발주처코드', '발주처', 
    '배송코드', '배송지', '상품코드', '상품명', 'UNIT수량', 
    'UNIT단가', '금        액', '부  가   세', 'LOT', '특이사항', 'Type', '특이사항'
]

# --- 2. 메인 UI ---
st.title("📂 Tesco 수주 데이터 처리 (FLOW 변환 포함)")
st.info("M, N, X열을 직접 삭제한 ordview 파일을 업로드해 주세요.")

# 서식 파일 존재 여부 확인 (오류 방지)
if not os.path.exists(MASTER_TEMPLATE):
    st.error(f"❌ 파일을 찾을 수 없음: '{MASTER_TEMPLATE}'이 서버(GitHub)에 업로드되어 있는지 확인하세요.")
else:
    uploaded_file = st.file_uploader("가공된 ordview 파일을 선택하세요", type=['xlsx', 'xls'])

    if uploaded_file:
        try:
            with st.spinner("데이터 변환 및 엑셀 수식 계산 중..."):
                # (1) 업로드된 데이터 읽기
                # 혹시 모를 헤더 오류를 방지하기 위해 0번 행부터 읽음
                df_input = pd.read_excel(uploaded_file)
                
                # (2) 'HYPER_FLOW' -> 'FLOW' 변환 (문자열인 경우에만 실행)
                df_input = df_input.astype(str).replace('HYPER_FLOW', 'FLOW', regex=True)
                
                # (3) 서식 파일 로드
                wb = openpyxl.load_workbook(MASTER_TEMPLATE)
                
                if '원본 데이터' not in wb.sheetnames:
                    st.error("❌ 서식 파일에 '원본 데이터' 시트가 없습니다.")
                else:
                    ws_source = wb['원본 데이터']
                    
                    # 기존 데이터 삭제 (2행부터 끝까지)
                    if ws_source.max_row > 1:
                        ws_source.delete_rows(2, ws_source.max_row)
                    
                    # 변환된 데이터 쓰기 (2행부터)
                    for r_idx, row in enumerate(df_input.values, 2):
                        for c_idx, value in enumerate(row, 1):
                            # 숫자로 변환 가능한 데이터는 숫자로 넣어줘야 엑셀 수식이 작동함
                            try:
                                if "." in str(value):
                                    val = float(value)
                                else:
                                    val = int(value)
                                ws_source.cell(row=r_idx, column=c_idx, value=val)
                            except:
                                ws_source.cell(row=r_idx, column=c_idx, value=value)
                
                # (4) 엑셀 내부 수식 결과 추출을 위한 임시 저장 및 재로드
                tmp_buffer = io.BytesIO()
                wb.save(tmp_buffer)
                tmp_buffer.seek(0)
                
                # data_only=True로 로드 (수식의 결과값만 가져옴)
                wb_eval = openpyxl.load_workbook(tmp_buffer, data_only=True)
                
                if 'Summary' not in wb_eval.sheetnames:
                    st.error("❌ 서식 파일에 'Summary' 시트가 없습니다.")
                else:
                    ws_summary = wb_eval['Summary']
                    
                    # Summary 시트 데이터프레임 변환
                    summary_data = ws_summary.values
                    summary_cols = next(summary_data)
                    df_summary = pd.DataFrame(summary_data, columns=summary_cols)
                    
                    # (5) 수량 0 제외 필터링
                    # '수량' 컬럼 이름이 정확한지 확인 필요 (Summary 시트 기준)
                    if '수량' in df_summary.columns:
                        df_summary['수량'] = pd.to_numeric(df_summary['수량'], errors='coerce').fillna(0)
                        df_valid = df_summary[df_summary['수량'] > 0].copy()
                    else:
                        st.error("❌ Summary 시트에 '수량' 컬럼을 찾을 수 없습니다.")
                        df_valid = pd.DataFrame()

                    if not df_valid.empty:
                        # 오늘 날짜
                        kst = timezone(timedelta(hours=9))
                        today = datetime.now(kst).strftime("%Y%m%d")
                        
                        # 통합 수주업로드 데이터 구성
                        res_df = pd.DataFrame(columns=FINAL_COLUMNS)
                        res_df['출고구분'] = 0
                        res_df['수주일자'] = today
                        res_df['납품일자'] = today 
                        res_df['발주처코드'] = df_valid.get('발주코드', '').fillna('')
                        res_df['발주처'] = "홈플러스"
                        res_df['배송코드'] = df_valid.get('배송코드', '').fillna('')
                        res_df['상품코드'] = df_valid.get('상품코드', '').fillna('')
                        res_df['상품명'] = df_valid.get('상품명', '').fillna('')
                        res_df['UNIT수량'] = df_valid['수량'].astype(int)
                        
                        # 단가 및 금액 계산
                        u_price = pd.to_numeric(df_valid.get('UNIT단가', 0), errors='coerce').fillna(0)
                        res_df['UNIT단가'] = u_price.astype(int)
                        res_df['금        액'] = (res_df['UNIT수량'] * res_df['UNIT단가']).astype(int)
                        res_df['부  가   세'] = (res_df['금        액'] * 0.1).astype(int)
                        res_df.fillna('', inplace=True)

                        # 화면 표시
                        st.success("✅ 변환 완료!")
                        st.dataframe(res_df, use_container_width=True)
                        
                        # 다운로드 파일 생성
                        final_out = io.BytesIO()
                        with pd.ExcelWriter(final_out, engine='xlsxwriter') as writer:
                            export_df = res_df.copy()
                            export_df.columns = EXPORT_HEADERS
                            export_df.to_excel(writer, index=False, sheet_name='서식')
                        
                        st.download_button(
                            label="📥 통합 수주업로드 파일 다운로드",
                            data=final_out.getvalue(),
                            file_name=f"통합수주_{today}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                    else:
                        st.warning("⚠️ 추출된 데이터가 없습니다. Summary 시트의 수식을 확인해 주세요.")

        except Exception as e:
            st.error(f"🚨 애플리케이션 실행 중 오류가 발생했습니다: {e}")
            st.info("해결 방법: 1. 서식파일의 시트 이름 확인 / 2. 업로드 파일의 열 순서 확인")
